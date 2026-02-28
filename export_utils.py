"""
Export Utilities
Generate Excel and PDF reports from report_engine output dicts.
"""

from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from fpdf import FPDF

EXPORTS_DIR = Path(__file__).parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

SJU_RED   = "CC0000"
SJU_WHITE = "FFFFFF"
DARK_GRAY = "1A1A1A"
LIGHT_GRAY= "F5F5F5"
MID_GRAY  = "CCCCCC"
GREEN     = "00AA44"
ORANGE    = "FF8C00"

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _hdr_fill(color: str = SJU_RED) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def _border() -> Border:
    thin = Side(style="thin", color=MID_GRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _write_header(ws, row: int, col: int, text: str, width_cols: int = 1,
                  bg: str = SJU_RED, fg: str = SJU_WHITE, size: int = 10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = _hdr_fill(bg)
    cell.font = _font(bold=True, color=fg, size=size)
    cell.alignment = _center()
    cell.border = _border()
    if width_cols > 1:
        end_col = get_column_letter(col + width_cols - 1)
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + width_cols - 1
        )
    return cell

def _write_cell(ws, row: int, col: int, value, bold=False, bg=None,
                align="center", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold)
    cell.alignment = _center() if align == "center" else _left()
    cell.border = _border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if number_format:
        cell.number_format = number_format
    return cell

def _rank_color(rank: int) -> str:
    if rank <= 25:
        return "C8F7C5"   # light green
    elif rank <= 75:
        return "FEFBD8"   # light yellow
    elif rank <= 150:
        return "FFE0B2"   # light orange
    else:
        return "FFCDD2"   # light red

def _autofit(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


# ─────────────────────────────────────────────────────────────
# SCOUT REPORT — Excel
# ─────────────────────────────────────────────────────────────

def export_scout_excel(report: dict) -> str:
    """Export pre-game scout report to Excel. Returns file path."""
    home = report["home_team"]
    away = report["away_team"]
    date = report["game_date"]
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    fname = EXPORTS_DIR / f"scout_{away}_vs_{home}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Overview ──────────────────────────────────────
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"PRE-GAME SCOUT REPORT  |  {away.upper()} at {home.upper()}  |  {date}"
    title.font = Font(bold=True, size=14, color=SJU_WHITE, name="Calibri")
    title.fill = _hdr_fill(DARK_GRAY)
    title.alignment = _center()
    ws.row_dimensions[1].height = 28

    r = 3
    # Column headers
    for col, txt in enumerate(["Metric", "St. John's", "Rank", "Opponent", "Rank", "Edge"], 1):
        _write_header(ws, r, col, txt)
    r += 1

    # Determine SJU side
    sju_ov  = report["home_overview"] if "john" in home.lower() else report["away_overview"]
    opp_ov  = report["away_overview"] if "john" in home.lower() else report["home_overview"]
    sju_ff  = report["home_ff"]       if "john" in home.lower() else report["away_ff"]
    opp_ff  = report["away_ff"]       if "john" in home.lower() else report["home_ff"]
    sju_sh  = report["home_shooting"] if "john" in home.lower() else report["away_shooting"]
    opp_sh  = report["away_shooting"] if "john" in home.lower() else report["home_shooting"]
    sju_ht  = report["home_height"]   if "john" in home.lower() else report["away_height"]
    opp_ht  = report["away_height"]   if "john" in home.lower() else report["home_height"]

    def edge(sju_val, opp_val, higher=True):
        if higher:
            return "SJU ▲" if sju_val > opp_val else "OPP ▲" if opp_val > sju_val else "EVEN"
        return "SJU ▲" if sju_val < opp_val else "OPP ▲" if opp_val < sju_val else "EVEN"

    def edge_color(e):
        return "C8F7C5" if e.startswith("SJU") else "FFCDD2" if e.startswith("OPP") else LIGHT_GRAY

    overview_rows = [
        ("Record",     sju_ov["record"], "",                   opp_ov["record"], "", ""),
        ("Adj EM",     f"{sju_ov['adj_em']:.2f}", _rank_color(sju_ov['rank_em']),
                       f"{opp_ov['adj_em']:.2f}", _rank_color(opp_ov['rank_em']),
                       edge(sju_ov['adj_em'], opp_ov['adj_em'])),
        ("Adj OE",     f"{sju_ov['adj_oe']:.1f}", _rank_color(sju_ov['rank_oe']),
                       f"{opp_ov['adj_oe']:.1f}", _rank_color(opp_ov['rank_oe']),
                       edge(sju_ov['adj_oe'], opp_ov['adj_oe'])),
        ("Adj DE",     f"{sju_ov['adj_de']:.1f}", _rank_color(sju_ov['rank_de']),
                       f"{opp_ov['adj_de']:.1f}", _rank_color(opp_ov['rank_de']),
                       edge(sju_ov['adj_de'], opp_ov['adj_de'], higher=False)),
        ("Tempo",      f"{sju_ov['tempo']:.1f}", "",
                       f"{opp_ov['tempo']:.1f}", "", ""),
        ("SOS",        f"{sju_ov['sos']:.3f}", _rank_color(sju_ov['rank_sos']),
                       f"{opp_ov['sos']:.3f}", _rank_color(opp_ov['rank_sos']), ""),
        # Four Factors
        ("",) * 6,
        ("— FOUR FACTORS (Offense) —",) + ("",) * 5,
        ("eFG%",       f"{sju_ff['efg_pct']:.1f}%", _rank_color(sju_ff['rank_efg']),
                       f"{opp_ff['efg_pct']:.1f}%", _rank_color(opp_ff['rank_efg']),
                       edge(sju_ff['efg_pct'], opp_ff['efg_pct'])),
        ("TO%",        f"{sju_ff['to_pct']:.1f}%", _rank_color(sju_ff['rank_to']),
                       f"{opp_ff['to_pct']:.1f}%", _rank_color(opp_ff['rank_to']),
                       edge(sju_ff['to_pct'], opp_ff['to_pct'], higher=False)),
        ("OR%",        f"{sju_ff['or_pct']:.1f}%", _rank_color(sju_ff['rank_or']),
                       f"{opp_ff['or_pct']:.1f}%", _rank_color(opp_ff['rank_or']),
                       edge(sju_ff['or_pct'], opp_ff['or_pct'])),
        ("FT Rate",    f"{sju_ff['ft_rate']:.3f}", _rank_color(sju_ff['rank_ft']),
                       f"{opp_ff['ft_rate']:.3f}", _rank_color(opp_ff['rank_ft']),
                       edge(sju_ff['ft_rate'], opp_ff['ft_rate'])),
        ("",) * 6,
        ("— FOUR FACTORS (Defense) —",) + ("",) * 5,
        ("Opp eFG%",   f"{sju_ff['d_efg']:.1f}%", _rank_color(sju_ff['rank_defg']),
                       f"{opp_ff['d_efg']:.1f}%", _rank_color(opp_ff['rank_defg']),
                       edge(sju_ff['d_efg'], opp_ff['d_efg'], higher=False)),
        ("Opp TO%",    f"{sju_ff['d_to']:.1f}%", _rank_color(sju_ff['rank_dto']),
                       f"{opp_ff['d_to']:.1f}%", _rank_color(opp_ff['rank_dto']),
                       edge(sju_ff['d_to'], opp_ff['d_to'])),
        ("Opp OR%",    f"{sju_ff['d_or']:.1f}%", _rank_color(sju_ff['rank_dor']),
                       f"{opp_ff['d_or']:.1f}%", _rank_color(opp_ff['rank_dor']),
                       edge(sju_ff['d_or'], opp_ff['d_or'], higher=False)),
        ("— SHOOTING —",) + ("",) * 5,
        ("3PT%",       f"{sju_sh['fg3_pct']:.1f}%", _rank_color(sju_sh['rank_fg3']),
                       f"{opp_sh['fg3_pct']:.1f}%", _rank_color(opp_sh['rank_fg3']),
                       edge(sju_sh['fg3_pct'], opp_sh['fg3_pct'])),
        ("2PT%",       f"{sju_sh['fg2_pct']:.1f}%", _rank_color(sju_sh['rank_fg2']),
                       f"{opp_sh['fg2_pct']:.1f}%", _rank_color(opp_sh['rank_fg2']),
                       edge(sju_sh['fg2_pct'], opp_sh['fg2_pct'])),
        ("FT%",        f"{sju_sh['ft_pct']:.1f}%", _rank_color(sju_sh['rank_ft']),
                       f"{opp_sh['ft_pct']:.1f}%", _rank_color(opp_sh['rank_ft']),
                       edge(sju_sh['ft_pct'], opp_sh['rank_ft'])),
        ("3PA Rate",   f"{sju_sh['f3g_rate']:.2f}", "",
                       f"{opp_sh['f3g_rate']:.2f}", "", ""),
        ("Assist Rate",f"{sju_sh['assist_rate']:.2f}", _rank_color(sju_sh['rank_assist']),
                       f"{opp_sh['assist_rate']:.2f}", _rank_color(opp_sh['rank_assist']),
                       edge(sju_sh['assist_rate'], opp_sh['assist_rate'])),
        ("Block%",     f"{sju_sh['block_pct']:.2f}", "",
                       f"{opp_sh['block_pct']:.2f}", "", ""),
        ("Steal Rate", f"{sju_sh['steal_rate']:.2f}", _rank_color(sju_sh['rank_steal']),
                       f"{opp_sh['steal_rate']:.2f}", _rank_color(opp_sh['rank_steal']),
                       edge(sju_sh['steal_rate'], opp_sh['steal_rate'])),
    ]

    for data_row in overview_rows:
        if len(data_row) == 1 and data_row[0] == "":
            r += 1
            continue
        metric = data_row[0]
        if metric.startswith("—"):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            cell = ws.cell(row=r, column=1, value=metric)
            cell.font = _font(bold=True, color=SJU_WHITE, size=10)
            cell.fill = _hdr_fill(SJU_RED)
            cell.alignment = _center()
            r += 1
            continue

        vals = data_row
        # metric
        c = ws.cell(row=r, column=1, value=vals[0])
        c.font = _font(bold=True)
        c.alignment = _left()
        c.border = _border()

        # sju value
        c2 = ws.cell(row=r, column=2, value=vals[1])
        c2.alignment = _center()
        c2.border = _border()
        if vals[2]:
            c2.fill = PatternFill("solid", fgColor=vals[2])

        # sju rank (from overview)
        sju_rank_map = {
            "Adj EM": sju_ov['rank_em'], "Adj OE": sju_ov['rank_oe'],
            "Adj DE": sju_ov['rank_de'], "SOS": sju_ov['rank_sos'],
            "eFG%": sju_ff['rank_efg'], "TO%": sju_ff['rank_to'],
            "OR%": sju_ff['rank_or'], "FT Rate": sju_ff['rank_ft'],
            "Opp eFG%": sju_ff['rank_defg'], "Opp TO%": sju_ff['rank_dto'],
            "Opp OR%": sju_ff['rank_dor'],
            "3PT%": sju_sh['rank_fg3'], "2PT%": sju_sh['rank_fg2'],
            "Assist Rate": sju_sh['rank_assist'], "Steal Rate": sju_sh['rank_steal'],
        }
        opp_rank_map = {
            "Adj EM": opp_ov['rank_em'], "Adj OE": opp_ov['rank_oe'],
            "Adj DE": opp_ov['rank_de'], "SOS": opp_ov['rank_sos'],
            "eFG%": opp_ff['rank_efg'], "TO%": opp_ff['rank_to'],
            "OR%": opp_ff['rank_or'], "FT Rate": opp_ff['rank_ft'],
            "Opp eFG%": opp_ff['rank_defg'], "Opp TO%": opp_ff['rank_dto'],
            "Opp OR%": opp_ff['rank_dor'],
            "3PT%": opp_sh['rank_fg3'], "2PT%": opp_sh['rank_fg2'],
            "Assist Rate": opp_sh['rank_assist'], "Steal Rate": opp_sh['rank_steal'],
        }
        sju_rank = sju_rank_map.get(vals[0], "")
        opp_rank = opp_rank_map.get(vals[0], "")
        c3 = ws.cell(row=r, column=3, value=f"#{sju_rank}" if sju_rank else "")
        c3.alignment = _center(); c3.border = _border()
        c4 = ws.cell(row=r, column=4, value=vals[3])
        c4.alignment = _center(); c4.border = _border()
        if vals[2]:
            c4.fill = PatternFill("solid", fgColor=vals[2])
        c5 = ws.cell(row=r, column=5, value=f"#{opp_rank}" if opp_rank else "")
        c5.alignment = _center(); c5.border = _border()
        # edge
        edge_val = vals[5] if len(vals) > 5 else ""
        c6 = ws.cell(row=r, column=6, value=edge_val)
        c6.alignment = _center(); c6.border = _border()
        c6.font = _font(bold=True)
        if edge_val.startswith("SJU"):
            c6.fill = PatternFill("solid", fgColor="C8F7C5")
        elif edge_val.startswith("OPP"):
            c6.fill = PatternFill("solid", fgColor="FFCDD2")

        r += 1

    _autofit(ws)

    # ── Sheet 2: Coaching Callouts ─────────────────────────────
    ws2 = wb.create_sheet("Coaching Callouts")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = f"COACHING CALLOUTS  |  {away.upper()} at {home.upper()}"
    t.font = Font(bold=True, size=13, color=SJU_WHITE, name="Calibri")
    t.fill = _hdr_fill(DARK_GRAY)
    t.alignment = _center()
    ws2.row_dimensions[1].height = 26

    r2 = 3
    priority_colors = {1: "FFCDD2", 2: "FFE0B2", 3: "FEFBD8"}
    priority_labels = {1: "🔴 CRITICAL", 2: "🟠 IMPORTANT", 3: "🟡 NOTE"}

    for callout in report.get("callouts", []):
        p = callout["priority"]
        bg = priority_colors.get(p, LIGHT_GRAY)

        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
        c = ws2.cell(row=r2, column=1, value=f"{priority_labels[p]}  |  {callout['label']}")
        c.font = _font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = _left()
        c.border = _border()
        ws2.row_dimensions[r2].height = 20
        r2 += 1

        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=3)
        c2 = ws2.cell(row=r2, column=1, value=callout["detail"])
        c2.font = _font(size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = _border()
        ws2.row_dimensions[r2].height = 40
        r2 += 2

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 60

    # ── Sheet 3: Win Probability ───────────────────────────────
    ws3 = wb.create_sheet("Win Probability")
    ws3.sheet_view.showGridLines = False
    fm = report.get("fanmatch", {})
    if fm:
        ws3.merge_cells("A1:D1")
        t3 = ws3["A1"]
        t3.value = "KENPOM WIN PROBABILITY"
        t3.font = Font(bold=True, size=13, color=SJU_WHITE, name="Calibri")
        t3.fill = _hdr_fill(DARK_GRAY)
        t3.alignment = _center()
        ws3.row_dimensions[1].height = 26

        rows3 = [
            ("Home Team", home, "Away Team", away),
            ("Predicted Score", f"{fm.get('home_pred', 0):.1f}",
             "Predicted Score", f"{fm.get('visitor_pred', 0):.1f}"),
            ("Win Probability", f"{fm.get('home_wp', 0)*100:.1f}%",
             "Win Probability", f"{(1 - fm.get('home_wp', 0))*100:.1f}%"),
            ("Predicted Tempo", f"{fm.get('pred_tempo', 0):.1f}", "", ""),
            ("Thrill Score", f"{fm.get('thrill', 0):.2f}", "", ""),
        ]
        for i, row_data in enumerate(rows3, start=3):
            for j, val in enumerate(row_data, start=1):
                c = ws3.cell(row=i, column=j, value=val)
                c.border = _border()
                c.alignment = _center()
                if j in (1, 3):
                    c.font = _font(bold=True)
                    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    wb.save(fname)
    return str(fname)


# ─────────────────────────────────────────────────────────────
# POST-GAME REPORT — Excel
# ─────────────────────────────────────────────────────────────

def export_postgame_excel(report: dict) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fname = EXPORTS_DIR / f"postgame_{report['sju_team']}_{ts}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Four Factors
    ws = wb.create_sheet("Four Factors")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    result_str = "WIN" if report["result"] == "W" else "LOSS"
    t.value = f"POST-GAME ANALYSIS  |  {report['sju_team']} {report['sju_score']} — {report['opp_score']} {report['opp_team']}  ({result_str})"
    t.font = Font(bold=True, size=13, color=SJU_WHITE, name="Calibri")
    t.fill = _hdr_fill("006400" if report["result"] == "W" else SJU_RED)
    t.alignment = _center()
    ws.row_dimensions[1].height = 26

    r = 3
    for col, hdr in enumerate(["Factor", "SJU", "Opponent", "Winner", "Details"], 1):
        _write_header(ws, r, col, hdr)
    r += 1

    for factor, (winner, detail) in report["ff_battle"].items():
        _write_cell(ws, r, 1, factor, bold=True, align="left")
        sju_ff = report["sju_ff"]
        opp_ff = report["opp_ff"]
        ff_map = {
            "eFG%": (f"{sju_ff['efg_pct']:.1f}%", f"{opp_ff['efg_pct']:.1f}%"),
            "TO%":  (f"{sju_ff['to_pct']:.1f}%",  f"{opp_ff['to_pct']:.1f}%"),
            "OR%":  (f"{sju_ff['or_pct']:.1f}%",  f"{opp_ff['or_pct']:.1f}%"),
            "FT Rate":(f"{sju_ff['ft_rate']:.3f}", f"{opp_ff['ft_rate']:.3f}"),
        }
        sv, ov = ff_map.get(factor, ("", ""))
        _write_cell(ws, r, 2, sv)
        _write_cell(ws, r, 3, ov)
        c_win = ws.cell(row=r, column=4, value=winner)
        c_win.alignment = _center(); c_win.border = _border()
        c_win.font = _font(bold=True)
        c_win.fill = PatternFill("solid", fgColor="C8F7C5" if "SJU" in winner else "FFCDD2")
        _write_cell(ws, r, 5, detail, align="left")
        r += 1

    # Narrative
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value="GAME NARRATIVE")
    c.font = _font(bold=True, color=SJU_WHITE, size=11)
    c.fill = _hdr_fill(DARK_GRAY)
    c.alignment = _center()
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c2 = ws.cell(row=r, column=1, value=report.get("narrative", ""))
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = _border()
    ws.row_dimensions[r].height = 60
    _autofit(ws)

    # Sheet 2: Player Grades
    grades = report.get("player_grades")
    if grades is not None and not grades.empty:
        ws2 = wb.create_sheet("Player Grades")
        ws2.sheet_view.showGridLines = False
        cols = list(grades.columns)
        for j, col_name in enumerate(cols, 1):
            _write_header(ws2, 1, j, col_name)
        grade_colors = {"A+":"C8F7C5","A":"D4EDDA","B":"FEFBD8","C":"FFE0B2","D":"FFCDD2","F":"F5C6CB"}
        for i, (_, row_data) in enumerate(grades.iterrows(), start=2):
            for j, val in enumerate(row_data.values, start=1):
                c = ws2.cell(row=i, column=j, value=str(val))
                c.alignment = _center(); c.border = _border()
                if cols[j-1] == "Grade":
                    c.fill = PatternFill("solid", fgColor=grade_colors.get(str(val), LIGHT_GRAY))
                    c.font = _font(bold=True)
        _autofit(ws2)

    wb.save(fname)
    return str(fname)


# ─────────────────────────────────────────────────────────────
# SEASON REPORT — Excel
# ─────────────────────────────────────────────────────────────

def export_season_excel(report: dict) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    tp = report["team_profile"]
    fname = EXPORTS_DIR / f"season_{tp['team'].replace(' ', '_')}_{ts}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Team Profile
    ws = wb.create_sheet("Team Profile")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"SEASON INTELLIGENCE  |  {tp['team'].upper()}  |  {tp['record']}  |  {tp['conf']}"
    t.font = Font(bold=True, size=13, color=SJU_WHITE, name="Calibri")
    t.fill = _hdr_fill(DARK_GRAY)
    t.alignment = _center()
    ws.row_dimensions[1].height = 26

    r = 3
    metrics = [
        ("Coach", tp["coach"]),
        ("Record", tp["record"]),
        ("Conference", tp["conf"]),
        ("Adj Efficiency Margin", f"{tp['adj_em']:.2f}  (#{tp['rank_em']} nationally / {tp['pctile_em']}th pctile)"),
        ("Adj Offensive Efficiency", f"{tp['adj_oe']:.1f}  (#{tp['rank_oe']} / {tp['pctile_oe']}th pctile)"),
        ("Adj Defensive Efficiency", f"{tp['adj_de']:.1f}  (#{tp['rank_de']} / {tp['pctile_de']}th pctile)"),
        ("Adjusted Tempo", f"{tp['tempo']:.1f} poss/40min  (#{tp['rank_tempo']})"),
        ("Strength of Schedule", f"{tp['sos']:.3f}  (#{tp['rank_sos']})"),
        ("Luck Rating", f"{tp['luck']:.3f}"),
        ("Team Experience", f"{tp['experience']:.2f}"),
        ("Bench Strength", f"{tp['bench']:.2f}"),
        ("Roster Continuity", f"{tp['continuity']:.2f}"),
    ]
    for label, value in metrics:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _font(bold=True)
        c1.alignment = _left()
        c1.border = _border()
        c1.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.alignment = _left()
        c2.border = _border()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # Sheet 2: Four Factors
    ws2 = wb.create_sheet("Four Factors")
    ws2.sheet_view.showGridLines = False
    ff = report["ff_profile"]
    ws2.merge_cells("A1:D1")
    t2 = ws2["A1"]
    t2.value = "FOUR FACTORS — SEASON AVERAGES"
    t2.font = Font(bold=True, size=12, color=SJU_WHITE, name="Calibri")
    t2.fill = _hdr_fill(SJU_RED)
    t2.alignment = _center()
    for col, hdr in enumerate(["Factor", "Value", "Rank", "Context"], 1):
        _write_header(ws2, 2, col, hdr)

    ff_rows = [
        ("OFFENSE", "", "", ""),
        ("eFG%",       f"{ff['efg_pct']:.1f}%", f"#{ff['rank_efg']}", "Effective shooting efficiency (3s worth 1.5x)"),
        ("TO%",        f"{ff['to_pct']:.1f}%",  f"#{ff['rank_to']}",  "Lower = better (fewer turnovers per possession)"),
        ("OR%",        f"{ff['or_pct']:.1f}%",  f"#{ff['rank_or']}",  "Offensive rebound rate"),
        ("FT Rate",    f"{ff['ft_rate']:.3f}",  f"#{ff['rank_ft']}",  "FTM / FGA — aggressiveness getting to line"),
        ("DEFENSE", "", "", ""),
        ("Opp eFG%",   f"{ff['d_efg']:.1f}%",  f"#{ff['rank_defg']}", "Lower = better"),
        ("Forced TO%", f"{ff['d_to']:.1f}%",   f"#{ff['rank_dto']}",  "Higher = more turnovers forced"),
        ("Opp OR%",    f"{ff['d_or']:.1f}%",   f"#{ff['rank_dor']}",  "Lower = better defensive rebounding"),
        ("FT Rate Allowed", f"{ff['d_ft']:.3f}",f"#{ff['rank_dft']}", "Lower = better"),
    ]
    for i, (factor, val, rank, ctx) in enumerate(ff_rows, start=3):
        is_section = factor in ("OFFENSE", "DEFENSE")
        if is_section:
            ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
            c = ws2.cell(row=i, column=1, value=factor)
            c.font = _font(bold=True, color=SJU_WHITE)
            c.fill = _hdr_fill(SJU_RED)
            c.alignment = _center()
        else:
            for j, v in enumerate([factor, val, rank, ctx], 1):
                c = ws2.cell(row=i, column=j, value=v)
                c.border = _border()
                c.alignment = _center() if j < 4 else _left()
                if j == 1:
                    c.font = _font(bold=True)
                    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                if j == 3 and rank:
                    try:
                        c.fill = PatternFill("solid", fgColor=_rank_color(int(rank.replace("#", ""))))
                    except Exception:
                        pass
    _autofit(ws2)

    # Sheet 3: Conference Standings
    ws3 = wb.create_sheet("Conference Standings")
    ws3.sheet_view.showGridLines = False
    conf_teams = report.get("conf_teams", pd.DataFrame())
    if not conf_teams.empty:
        ws3.merge_cells("A1:F1")
        t3 = ws3["A1"]
        t3.value = f"BIG EAST STANDINGS — NATIONAL EFFICIENCY RANKINGS"
        t3.font = Font(bold=True, size=12, color=SJU_WHITE, name="Calibri")
        t3.fill = _hdr_fill(DARK_GRAY)
        t3.alignment = _center()
        hdrs = ["Team", "AdjEM", "Rank", "AdjOE", "AdjDE", "Tempo"]
        for j, h in enumerate(hdrs, 1):
            _write_header(ws3, 2, j, h)
        display_cols = ["TeamName", "AdjEM", "RankAdjEM", "AdjOE", "AdjDE", "AdjTempo"]
        available = [c for c in display_cols if c in conf_teams.columns]
        for i, (_, row_data) in enumerate(conf_teams[available].iterrows(), start=3):
            for j, val in enumerate(row_data.values, start=1):
                c = ws3.cell(row=i, column=j, value=val)
                c.border = _border()
                c.alignment = _center() if j > 1 else _left()
                if str(row_data.get("TeamName", "")).lower() == tp["team"].lower():
                    c.font = _font(bold=True, color=SJU_WHITE)
                    c.fill = _hdr_fill(SJU_RED)
        _autofit(ws3)

    wb.save(fname)
    return str(fname)


# ─────────────────────────────────────────────────────────────
# Simple PDF wrapper (uses fpdf2)
# ─────────────────────────────────────────────────────────────

def _pdf_safe(text: str) -> str:
    """Remove characters outside Latin-1 range (e.g. emojis) for Helvetica compatibility."""
    return text.encode("latin-1", errors="ignore").decode("latin-1")

class _PDF(FPDF):
    def header(self):
        self.set_fill_color(26, 26, 26)
        self.rect(0, 0, 210, 14, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 12)
        self.set_y(3)
        self.cell(0, 8, "RedStorm Analytics Hub", align="C")
        self.set_text_color(0, 0, 0)
        self.set_xy(self.l_margin, 18)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 6, f"Page {self.page_no()} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")


def export_scout_pdf(report: dict) -> str:
    """Export scout report to PDF."""
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    home = report["home_team"]
    away = report["away_team"]
    fname = EXPORTS_DIR / f"scout_{away}_vs_{home}_{ts}.pdf"

    pdf = _PDF()
    pdf.set_margins(left=10, top=18, right=10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_fill_color(204, 0, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"PRE-GAME SCOUT: {away.upper()} at {home.upper()}", fill=True, align="C")
    pdf.ln()
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Game Date: {report['game_date']}", align="C")
    pdf.ln()
    pdf.ln(1)
    pdf.set_text_color(0, 0, 0)

    def section(title):
        pdf.set_fill_color(26, 26, 26)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, f"  {title}", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 9)

    def row2(label, v1, v2):
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(65, 6, label, border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(60, 6, str(v1), border=1, align="C")
        pdf.cell(60, 6, str(v2), border=1, align="C")
        pdf.ln()

    sju_ov = report["home_overview"] if "john" in home.lower() else report["away_overview"]
    opp_ov = report["away_overview"] if "john" in home.lower() else report["home_overview"]

    # Header row
    pdf.set_fill_color(204, 0, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(65, 6, "METRIC", border=1, fill=True)
    pdf.cell(60, 6, "ST. JOHN'S", border=1, fill=True, align="C")
    pdf.cell(60, 6, "OPPONENT", border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    row2("Record", sju_ov["record"], opp_ov["record"])
    row2("Adj Efficiency Margin", f"{sju_ov['adj_em']:.2f} (#{sju_ov['rank_em']})", f"{opp_ov['adj_em']:.2f} (#{opp_ov['rank_em']})")
    row2("Adj Offense (AdjOE)", f"{sju_ov['adj_oe']:.1f} (#{sju_ov['rank_oe']})", f"{opp_ov['adj_oe']:.1f} (#{opp_ov['rank_oe']})")
    row2("Adj Defense (AdjDE)", f"{sju_ov['adj_de']:.1f} (#{sju_ov['rank_de']})", f"{opp_ov['adj_de']:.1f} (#{opp_ov['rank_de']})")
    row2("Tempo", f"{sju_ov['tempo']:.1f}", f"{opp_ov['tempo']:.1f}")

    pdf.ln(3)
    section("COACHING CALLOUTS")
    pdf.ln(2)
    priority_labels = {1: "[CRITICAL]", 2: "[IMPORTANT]", 3: "[NOTE]"}
    for callout in report.get("callouts", []):
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.w - pdf.l_margin - pdf.r_margin, 5, _pdf_safe(f"{priority_labels.get(callout['priority'], '')} {callout['label']}"))
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(60, 60, 60)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.w - pdf.l_margin - pdf.r_margin, 4, _pdf_safe(callout["detail"]))
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    fm = report.get("fanmatch", {})
    if fm:
        pdf.ln(3)
        section("WIN PROBABILITY (KenPom)")
        pdf.ln(2)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(pdf.w - pdf.l_margin - pdf.r_margin, 5,
            f"{home} Win Probability: {fm.get('home_wp', 0):.1f}%  |  "
            f"{away} Win Probability: {100 - fm.get('home_wp', 0):.1f}%\n"
            f"Predicted Score: {home} {fm.get('home_pred', 0):.0f} - {away} {fm.get('visitor_pred', 0):.0f}\n"
            f"Predicted Tempo: {fm.get('pred_tempo', 0):.1f} poss/40min"
        )

    pdf.output(str(fname))
    return str(fname)
